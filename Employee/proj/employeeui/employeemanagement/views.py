import psycopg2
import xlrd
from django.shortcuts import render, redirect
from django.http import HttpResponse
from employeemanagement.models import EmployeeManagement
from django.contrib import messages
# Create your views here.
from .fusioncharts import FusionCharts
from django.core.files.storage import FileSystemStorage
from rest_framework.response import Response
from openpyxl import Workbook
import xlwt
import openpyxl
# conn = psycopg2.connect(
#     database="employee_management",
#     user="employee",
#     password="iness",
#     host="localhost",
#     port="5432")
# cur = conn.cursor()


def home(request):
    return render(request, 'home.html', {})


def new_employee(request):
    print(request.method)
    if request.method == 'GET':
        return render(request, "new_employee.html")

    elif request.method == 'POST':
        name = request.POST.get("name")
        employee_code = request.POST.get("employee_code")
        email_id = request.POST.get("email_id")
        contact_no = request.POST.get("contact_no")
        salary = request.POST.get("salary")
        details = EmployeeManagement(
            name=name, employee_code=employee_code, email_id=email_id, contact_no=contact_no, salary=salary)
        details.save()
        return redirect('/newemployee')


def search_employee(request):
    if request.method == 'GET':
        return render(request, "search_employee.html")
    elif request.method == 'POST':
        name_to_search = request.POST.get("name")
        employees = EmployeeManagement.objects.filter(
            name__iregex=name_to_search)
    return render(request, "search_employee.html", context={"data": employees})


def show_employee(request):
    details = EmployeeManagement.objects.all()
    return render(request, 'show_employee.html', context={'details': details})


def delete_employee(request, id):
    details = EmployeeManagement.objects.get(id=id)
    details.delete()
    return redirect('/showemployee')


def edit_employee(request, id):
    if request.method == 'GET':
        employee = EmployeeManagement.objects.get(id=id)
        show = {
            "id": employee.id,
            "name": employee.name,
            "employee_code": employee.employee_code,
            "email_id": employee.email_id,
            "contact_no": employee.contact_no,
            "salary": employee.salary}

        return render(request, 'edit_employee.html', context=show)

    elif request.method == 'POST':
        email_id = request.POST.get("email_id")
        contact_no = request.POST.get("contact_no")
        salary = request.POST.get("salary")
        employee = EmployeeManagement.objects.get(id=id)
        employee.email_id = email_id
        employee.contact_no = contact_no
        employee.salary = salary
        employee.save()
        return redirect("/showemployee")


def file_upload(request):

    if "GET" == request.method:
        return render(request, 'file_upload.html', {})
    else:
        excel_file = request.FILES['excel_file']
        wb = openpyxl.load_workbook(excel_file)
        worksheet = wb["Sheet1"]
        print(worksheet)
        excel_data = list()
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data)
            detail = EmployeeManagement(
                name=row_data[0], employee_code=row_data[1], email_id=row_data[2], contact_no=row_data[3], salary=row_data[4])
            detail.save()
        details = EmployeeManagement.objects.all()

        return render(request, 'show_employee.html', context={'details': details})


def employee_reports(request):
    chartObj = FusionCharts('pie3d', 'ex1', '800', '600', 'chart-1', 'json', """{
  "chart": {
    "caption": "Employee Management",
    "showvalues": "1",
    "showpercentintooltip": "0",
    "enablemultislicing": "1",
    "theme": "fusion"
  },
  "data": [
    {
      "label": "Number of employee(10)",
      "value": "300000"
    },
    {
      "label": "Number of employee(10)",
      "value": "230000"
    },
    {
      "label": "Number of employee(10)",
      "value": "180000"
    },
    {
      "label": "Number of employee(15)",
      "value": "250000"
    },
    {
      "label": "Number of employee(5)",
      "value": "30000"
    }
  ]
}""")
    return render(request, 'employee_reports.html', {'output': chartObj.render()})


def export_xlsx(request):
    """
    Downloads all employee as Excel file with a single worksheet
    """
    employee = EmployeeManagement.objects.all()

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'employee'

    # Define the titles for columns
    columns = [
        'id',
        'name',
        'employee_code',
        'contact_no',
        'salary',
    ]
    row_num = 1
    # Assign the names for each cell of the header
    for col_num, column_name in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_name
    # Iterate through all movies
    for data in employee:
        row_num += 1
        # Define the data for each cell in the row
        row = [
            data.id,
            data.name,
            data.employee_code,
            data.contact_no,
            data.salary,
        ]
        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response
